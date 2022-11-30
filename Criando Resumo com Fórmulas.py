#Abrindo arquivo Excel e criando novas abas com as informações de cada funcionario
#Neste exemplo é necessário que o arquivo excel já esta preenchido na primeira página
#https://www.udemy.com/course/python-rpa-e-excel-aprenda-automatizar-processos-e-planilhas/learn/lecture/27889072#overview

from openpyxl import load_workbook
import os

nome_arquivo = "C:\\Users\\Windows\\Desktop\\Python Projetos\\openpyxl\\ExcelEmail\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#Seleciona a Sheet de Dados
sheet_selecionada = planilha_aberta['Dados']

nomeNovo = ""

#Para indicar a partir de qual linha devemos ler o arquivo
totalLinha = len(sheet_selecionada['A']) + 1

#Para tornar dinâmico a leitura das linhas, para ler todas que estiverem com informações
for linha in range(2, len(sheet_selecionada['A']) + 1):
    nomeAtual = sheet_selecionada['A%s' % linha].value
    if nomeNovo == nomeAtual:
        print(nomeAtual)

    else:
        #Cria uma nova sheet com o nome do funcionario
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)

        #Seleciona a sheet que foi criada
        sheet_selecionada2 = planilha_aberta[nomeAtual]

        #Adiciona o nome do funcionario que esta na linha que o código está passando
        nomeAtual = sheet_selecionada['A%s' % linha].value

        #Coloca os titulos
        sheet_selecionada2['A1'] = "Vendedor"
        sheet_selecionada2['B1'] = "Produtos"
        sheet_selecionada2['B1'] = "Vendas"

        #Preenche as informações na segunda linha
        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['C2'] = sheet_selecionada['C%s' % linha].value

#Salva as alterações feitas
planilha_aberta.save(filename=nome_arquivo)

#Abre a planilha
os.startfile(nome_arquivo)